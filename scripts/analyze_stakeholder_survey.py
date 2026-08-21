import json
import re
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE = Path(r"D:\my job\หลักสูตร\1.ปวส.พืชศาสตร์\แบบสอบถามความต้องการของผู้มีส่วนได้ส่วนเสียต่อสมรรถนะผู้สำเร็จการศึกษา หลักสูตรประกาศนียบัตรวิชาชีพชั้นสูง (ปวส.) สาขาวิชาพืชศาสตร์ (การตอบกลับ).xlsx")
OUTPUT = Path(__file__).with_name("stakeholder_survey_summary.json")


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def bracket_label(header):
    match = re.search(r"\[([^\]]+)\]\s*$", clean(header))
    return match.group(1).strip() if match else clean(header)


def stats(rows, columns):
    result = []
    for index in columns:
        values = [float(row[index]) for row in rows if isinstance(row[index], (int, float))]
        if not values:
            continue
        label = bracket_label(headers[index])
        result.append({
            "column": index + 1,
            "label": label,
            "n": len(values),
            "mean": round(sum(values) / len(values), 2),
            "high_45_n": sum(value >= 4 for value in values),
            "high_45_pct": round(100 * sum(value >= 4 for value in values) / len(values), 1),
            "distribution": {str(score): sum(value == score for value in values) for score in range(1, 6)},
        })
    return sorted(result, key=lambda item: (-item["mean"], -item["high_45_pct"], item["label"]))


def split_choices(value):
    if value is None:
        return []
    return [part.strip() for part in str(value).split(",") if part.strip()]


workbook = openpyxl.load_workbook(SOURCE, data_only=True, read_only=False)
sheet = workbook[workbook.sheetnames[0]]
all_rows = list(sheet.iter_rows(values_only=True))
headers = [clean(value) for value in all_rows[0]]
rows = [list(row) for row in all_rows[1:] if any(value is not None for value in row)]

timestamps = [row[0] for row in rows if isinstance(row[0], datetime)]
groups = Counter(clean(row[1]) or "ไม่ระบุ" for row in rows)
experience = Counter(clean(row[3]) or "ไม่ระบุ" for row in rows)

# Rating blocks: columns E:AN and AQ:BB in the source workbook.
current_ratings = stats(rows, range(4, 40))
future_ratings = stats(rows, range(42, 54))

top5_choices = Counter()
must_do = Counter()
for row in rows:
    top5_choices.update(split_choices(row[40]))
    value = clean(row[41])
    if value:
        must_do[value] += 1

multi_select_columns = {}
for index in range(56, len(headers)):
    counts = Counter()
    for row in rows:
        counts.update(split_choices(row[index]))
    if counts:
        multi_select_columns[str(index + 1)] = {
            "header": headers[index],
            "responses": sum(1 for row in rows if clean(row[index])),
            "top": [{"label": key, "count": count} for key, count in counts.most_common(12)],
        }

summary = {
    "source": SOURCE.name,
    "sheet": sheet.title,
    "responses": len(rows),
    "date_min": min(timestamps).isoformat(sep=" ") if timestamps else None,
    "date_max": max(timestamps).isoformat(sep=" ") if timestamps else None,
    "groups": [{"label": key, "count": count, "pct": round(100 * count / len(rows), 1)} for key, count in groups.most_common()],
    "experience": [{"label": key, "count": count, "pct": round(100 * count / len(rows), 1)} for key, count in experience.most_common()],
    "current_ratings": current_ratings,
    "future_ratings": future_ratings,
    "top5_priorities": [{"label": key, "count": count, "pct_respondents": round(100 * count / len(rows), 1)} for key, count in top5_choices.most_common()],
    "must_do": [{"label": key, "count": count, "pct_respondents": round(100 * count / len(rows), 1)} for key, count in must_do.most_common()],
    "multi_select": multi_select_columns,
}

OUTPUT.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
print(json.dumps({
    "responses": summary["responses"],
    "date_min": summary["date_min"],
    "date_max": summary["date_max"],
    "groups": summary["groups"],
    "top_current": current_ratings[:10],
    "top_future": future_ratings[:8],
    "top5_priorities": summary["top5_priorities"][:12],
    "must_do": summary["must_do"][:12],
}, ensure_ascii=False, indent=2))
